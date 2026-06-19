"""
Padronizador de base diária de credenciados — PUCPR / Grupo Marista
Lê todos os CSVs de ../bases/ e gera os XLSXs na mesma pasta.
Uso: python main.py
"""

import os
import glob
import datetime
import unicodedata
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paleta PUCPR ──────────────────────────────────────────────────────────────
PRIMARY_PURE     = "8A0538"
PRIMARY_DARK     = "570013"
PRIMARY_DARKEST  = "300000"
PRIMARY_LIGHTEST = "E5C3D0"
LIGHT_PURE_02    = "F0F2F2"
LIGHT_PURE       = "FFFFFF"
DARK_02          = "404040"

SKIP_KEYWORDS = ['Cartão', 'Credencial', 'Página', 'UNIDADE', 'Relatório',
                 'Emissão', 'Período', 'Usuário:', 'Cartão:']

HEADERS_PADRAO = ['CARTÃO', 'CREDENCIAL', 'DATA', 'EVENTO', 'USUÁRIO', 'DESCRIÇÃO', 'GRUPO']
LARGURAS_PADRAO = [12, 14, 20, 22, 44, 14, 24]
LARGURA_COLUNA_DINAMICA = 20

# ── Caminho da pasta de bases (relativo ao script) ────────────────────────────
PASTA_BASES = os.path.join(os.path.dirname(__file__), '..', 'bases')

# ── Consolidação histórica por unidade ────────────────────────────────────────
NOME_CONSOLIDADO = 'consolidado.xlsx'
ABA_CURITIBA     = 'PUCPR - CURITIBA'
ABA_TOLEDO       = 'PUCPR - TOLEDO'
CHAVE_DEDUP      = ['CARTÃO', 'CREDENCIAL', 'DATA', 'EVENTO']


def fill(hex_color):
    return PatternFill('solid', start_color=hex_color)


def _parse_csv(filepath):
    with open(filepath, 'r', encoding='ISO-8859-1') as f:
        lines = f.readlines()

    meta = {'unidade': '', 'emissao': '', 'periodo': '', 'usuario': ''}
    for line in lines[:7]:
        cols = line.strip().split(';')
        joined = ' '.join(c.strip() for c in cols if c.strip())
        if 'UNIDADE'   in joined: meta['unidade'] = joined.strip()
        elif 'Emissão'  in joined: meta['emissao'] = joined.replace('Emissão:', '').strip()
        elif 'Período'  in joined: meta['periodo'] = joined.replace('Período:', '').replace('até', 'ATÉ').strip()
        elif 'Usuário:' in joined: meta['usuario'] = joined.replace('Usuário:', '').strip()

    rows = []
    for line in lines:
        cols = line.strip().split(';')
        while len(cols) < 16:
            cols.append('')
        if any(kw in ';'.join(cols) for kw in SKIP_KEYWORDS):
            continue
        if not any(c.strip() for c in cols):
            continue

        cartao     = (cols[1].strip() or cols[2].strip()).upper()
        credencial = (cols[3].strip() if cols[1].strip() else cols[2].strip()).upper()
        data       = cols[4].strip().upper()
        evento     = cols[6].strip().upper()
        usuario    = cols[10].strip().upper()
        descricao  = cols[12].strip().upper()
        grupo      = cols[15].strip().upper()

        if data and '/' in data:
            rows.append([cartao, credencial, data, evento, usuario, descricao, grupo])

    return meta, rows


def _build_xlsx(meta, headers, rows, output_path, titulo='RELATÓRIO DE CARTÕES DE CREDENCIADOS'):
    wb = Workbook()
    ws = wb.active
    ws.title = os.path.splitext(os.path.basename(output_path))[0][:31]

    ultima_coluna = get_column_letter(len(headers))

    borda = Border(
        left=Side(style='thin', color='E4E4E4'),
        right=Side(style='thin', color='E4E4E4'),
        top=Side(style='thin', color='E4E4E4'),
        bottom=Side(style='thin', color='E4E4E4')
    )
    centro = Alignment(horizontal='center', vertical='center')
    esq    = Alignment(horizontal='left',   vertical='center')

    ws.merge_cells(f'A1:{ultima_coluna}1')
    ws['A1'] = meta['unidade'].upper()
    ws['A1'].font      = Font(name='Poppins', bold=True, size=13, color='FFFFFF')
    ws['A1'].fill      = fill(PRIMARY_PURE)
    ws['A1'].alignment = centro
    ws.row_dimensions[1].height = 24

    ws.merge_cells(f'A2:{ultima_coluna}2')
    ws['A2'] = titulo.upper()
    ws['A2'].font      = Font(name='Poppins', bold=True, size=11, color='FFFFFF')
    ws['A2'].fill      = fill(PRIMARY_DARK)
    ws['A2'].alignment = centro
    ws.row_dimensions[2].height = 20

    ws.merge_cells(f'A3:{ultima_coluna}3')
    ws['A3'] = (
        f"EMISSÃO: {meta['emissao'].upper()}"
        f"   |   PERÍODO: {meta['periodo'].upper()}"
        f"   |   USUÁRIO: {meta['usuario'].upper()}"
    )
    ws['A3'].font      = Font(name='Poppins', size=9, color=PRIMARY_LIGHTEST)
    ws['A3'].fill      = fill(PRIMARY_DARKEST)
    ws['A3'].alignment = centro
    ws.row_dimensions[3].height = 16

    ws.merge_cells(f'A4:{ultima_coluna}4')
    ws['A4'].fill = fill(LIGHT_PURE_02)
    ws.row_dimensions[4].height = 6

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=h)
        cell.font      = Font(name='Poppins', bold=True, size=10, color='FFFFFF')
        cell.fill      = fill(PRIMARY_PURE)
        cell.alignment = centro
        cell.border    = Border(bottom=Side(style='medium', color=PRIMARY_DARK))
    ws.row_dimensions[5].height = 20

    for i, row in enumerate(rows):
        excel_row = i + 6
        bg = LIGHT_PURE if i % 2 == 0 else LIGHT_PURE_02
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.font      = Font(name='Source Sans Pro', size=9, color=DARK_02)
            cell.fill      = fill(bg)
            cell.border    = borda
            cell.alignment = esq
        ws.row_dimensions[excel_row].height = 15

    larguras = LARGURAS_PADRAO if headers == HEADERS_PADRAO else [LARGURA_COLUNA_DINAMICA] * len(headers)
    for idx, largura in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = largura

    ws.auto_filter.ref = f"A5:{ultima_coluna}{len(rows) + 5}"
    ws.freeze_panes    = 'A6'

    wb.save(output_path)


def processar(csv_path):
    nome_base   = os.path.splitext(os.path.basename(csv_path))[0]
    output_path = os.path.join(os.path.dirname(csv_path), f"{nome_base}.xlsx")

    print(f"Processando: {os.path.basename(csv_path)}")
    meta, rows = _parse_csv(csv_path)
    _build_xlsx(meta, HEADERS_PADRAO, rows, output_path)
    print(f"  ✓ {len(rows)} registros → {os.path.basename(output_path)}")


# ── Ingestão de bases brutas em XLSX (ex.: relatórios de cancela/catraca) ─────
# O sistema institucional também pode exportar a base diretamente em XLSX, com
# layout próprio (diferente do CSV) e suas próprias colunas. Essas bases são
# padronizadas visualmente preservando todas as colunas originais, sem forçar
# o padrão de 7 colunas usado pelo CSV de cartões de credenciados.

def _ja_processado(ws):
    """Verifica se o XLSX já foi estilizado por este script (cor de fundo PUCPR em A1)."""
    cel = ws['A1']
    if not cel.fill or cel.fill.patternType != 'solid':
        return False
    cor = str(cel.fill.start_color.rgb or '').upper()
    return cor.endswith(PRIMARY_PURE.upper())


def _normalizar_valor(valor):
    if valor is None:
        return ''
    if isinstance(valor, datetime.datetime):
        return valor.strftime('%d/%m/%y %H:%M:%S')
    return str(valor).strip().upper()


def _extrair_meta_generica(ws, linha_cabecalho):
    """Lê as linhas de metadados (texto livre tipo 'Chave: valor') acima do cabeçalho."""
    meta_bruta = {}
    for r in range(1, linha_cabecalho):
        valores = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        texto = next((v for v in valores if isinstance(v, str) and v.strip()), None)
        if texto and ':' in texto:
            chave, _, valor = texto.partition(':')
            meta_bruta[chave.strip()] = valor.strip()
    return meta_bruta


def _ler_xlsx_bruto(xlsx_path):
    """Lê uma base bruta em XLSX (não gerada por este script): metadados em texto livre,
    cabeçalho e dados localizados dinamicamente, preservando todas as colunas originais."""
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    linha_cabecalho, indices = _detectar_cabecalho(ws)
    if linha_cabecalho is None:
        return None

    meta_bruta = _extrair_meta_generica(ws, linha_cabecalho)
    headers = [str(ws.cell(row=linha_cabecalho, column=c).value).strip() for c in indices]

    rows = []
    for r in range(linha_cabecalho + 1, ws.max_row + 1):
        valores = [_normalizar_valor(ws.cell(row=r, column=c).value) for c in indices]
        if any(v for v in valores):
            rows.append(valores)

    return meta_bruta, headers, rows


def obter_unidade_arquivo(texto_relatorio, nome_arquivo):
    """Identifica a unidade priorizando o texto dos metadados do relatório; usa o nome do
    arquivo como reserva quando o texto não traz a informação."""
    return obter_unidade(texto_relatorio) or obter_unidade(nome_arquivo)


def processar_xlsx_bruto(xlsx_path):
    """Padroniza visualmente uma base bruta em XLSX, preservando todas as colunas originais.
    Arquivos já estilizados por este script são ignorados (idempotência)."""
    wb_estilo = load_workbook(xlsx_path)
    if _ja_processado(wb_estilo.active):
        return False

    resultado = _ler_xlsx_bruto(xlsx_path)
    if resultado is None:
        print(f"  [AVISO] Cabeçalho não localizado em {os.path.basename(xlsx_path)} — ignorado")
        return False
    meta_bruta, headers, rows = resultado

    texto_relatorio = meta_bruta.get('Relatório', meta_bruta.get('Relatorio', ''))
    nome_aba = obter_unidade_arquivo(texto_relatorio, os.path.basename(xlsx_path))
    if nome_aba is None:
        print(f"  [AVISO] Unidade não identificada em {os.path.basename(xlsx_path)} — ignorado")
        return False

    meta = {
        'unidade': nome_aba,
        'emissao': meta_bruta.get('Data', ''),
        'periodo': meta_bruta.get('Carimbo de tempo do evento', ''),
        'usuario': meta_bruta.get('Usuário', meta_bruta.get('Usuario', '')),
    }
    titulo = texto_relatorio or 'RELATÓRIO'

    print(f"Processando (XLSX bruto): {os.path.basename(xlsx_path)}")
    _build_xlsx(meta, headers, rows, xlsx_path, titulo=titulo)
    print(f"  ✓ {len(rows)} registros → {os.path.basename(xlsx_path)} ({nome_aba})")
    return True


# ── Consolidação histórica por unidade ────────────────────────────────────────

def _normalizar(texto):
    """Remove acentos e caixa para comparação robusta (ex.: 'PARANÁ' -> 'PARANA')."""
    sem_acento = unicodedata.normalize('NFKD', texto or '')
    return ''.join(c for c in sem_acento if not unicodedata.combining(c)).upper()


def obter_unidade(texto_unidade):
    """Identifica a aba de destino (Curitiba/Toledo) a partir do texto de unidade do relatório."""
    texto = _normalizar(texto_unidade)
    if 'TOLEDO' in texto:
        return ABA_TOLEDO
    if 'PARANA' in texto:
        return ABA_CURITIBA
    return None


def _detectar_cabecalho(ws, limite_busca=15):
    """Procura, nas primeiras linhas, a primeira com 3+ células preenchidas — funciona tanto
    para os XLSX já estilizados (cabeçalho sempre com várias colunas separadas) quanto para
    bases brutas em XLSX (metadados em texto livre ocupam só 1 célula por linha, sem confundir
    com o cabeçalho real). Retorna (linha, lista de índices de coluna preenchidos)."""
    for r in range(1, limite_busca + 1):
        valores = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        indices = [c for c, v in enumerate(valores, start=1) if v not in (None, '')]
        if len(indices) >= 3:
            return r, indices
    return None, []


def _ler_individual(xlsx_path):
    """Lê unidade, cabeçalho e dados de um XLSX já gerado por _build_xlsx, localizando o
    cabeçalho dinamicamente para suportar colunas novas ou layouts ligeiramente diferentes."""
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    unidade_texto = ws['A1'].value or ''

    linha_cabecalho, indices = _detectar_cabecalho(ws)
    if linha_cabecalho is None:
        return unidade_texto, [], []

    headers = [str(ws.cell(row=linha_cabecalho, column=c).value).strip() for c in indices]

    rows = []
    for r in range(linha_cabecalho + 1, ws.max_row + 1):
        valores = [ws.cell(row=r, column=c).value for c in indices]
        valores = [v if v is not None else '' for v in valores]
        if any(str(v).strip() for v in valores):
            rows.append(valores)

    return unidade_texto, headers, rows


def carregar_consolidado(caminho):
    """Abre o consolidado existente ou cria um novo, garantindo as duas abas obrigatórias."""
    if os.path.exists(caminho):
        wb = load_workbook(caminho)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    for nome_aba in (ABA_CURITIBA, ABA_TOLEDO):
        if nome_aba not in wb.sheetnames:
            wb.create_sheet(nome_aba)

    return wb


def _garantir_cabecalho(ws, headers_origem):
    """Escreve o cabeçalho na aba se ainda não existir; acrescenta colunas novas ao final, se houver."""
    cabecalho_atual = [c.value for c in ws[1] if c.value] if ws['A1'].value else []

    novas_colunas = [h for h in headers_origem if h not in cabecalho_atual]
    inicio = len(cabecalho_atual) + 1
    for offset, h in enumerate(novas_colunas):
        cell = ws.cell(row=1, column=inicio + offset, value=h)
        cell.font = Font(name='Poppins', bold=True, size=10, color='FFFFFF')
        cell.fill = fill(PRIMARY_PURE)
    cabecalho_atual.extend(novas_colunas)

    if cabecalho_atual:
        ws.freeze_panes = 'A2'

    return cabecalho_atual


def registro_ja_existe(chave, chaves_existentes):
    return chave in chaves_existentes


def _indices_chave_dedup(cabecalho):
    """Localiza os índices de CARTÃO/CREDENCIAL/DATA/EVENTO de forma case/acento-insensitive
    (necessário pois formatos de origem diferentes, ex. relatório de cancelas, usam grafia
    própria). Se o formato não tiver os 4 campos padrão, usa a linha inteira como chave —
    garante deduplicação segura sem arriscar falso positivo entre registros distintos."""
    normalizados = [_normalizar(h) for h in cabecalho]
    chave_normalizada = [_normalizar(c) for c in CHAVE_DEDUP]
    if all(c in normalizados for c in chave_normalizada):
        return [normalizados.index(c) for c in chave_normalizada]
    return list(range(len(cabecalho)))


def inserir_registros(ws, cabecalho, headers_origem, rows):
    """Insere apenas registros inéditos abaixo do histórico já presente na aba."""
    indices_chave = _indices_chave_dedup(cabecalho)
    mapa_colunas  = [cabecalho.index(h) for h in headers_origem]

    chaves_existentes = set()
    for linha in ws.iter_rows(min_row=2, values_only=True):
        chave = tuple(str(linha[i]).strip().upper() if i < len(linha) and linha[i] is not None else ''
                       for i in indices_chave)
        chaves_existentes.add(chave)

    proxima_linha = ws.max_row + 1
    inseridos = 0
    for row in rows:
        linha_destino = [''] * len(cabecalho)
        for col_origem, col_destino in enumerate(mapa_colunas):
            linha_destino[col_destino] = row[col_origem]

        chave = tuple(str(linha_destino[i]).strip().upper() for i in indices_chave)
        if registro_ja_existe(chave, chaves_existentes):
            continue

        for col_idx, valor in enumerate(linha_destino, start=1):
            cell = ws.cell(row=proxima_linha, column=col_idx, value=valor)
            cell.font = Font(name='Source Sans Pro', size=9, color=DARK_02)

        chaves_existentes.add(chave)
        proxima_linha += 1
        inseridos += 1

    return inseridos


def consolidar_planilhas(pasta_bases):
    """Acumula no consolidado.xlsx os dados de todos os XLSX da pasta, separados por unidade."""
    caminho_consolidado = os.path.join(pasta_bases, NOME_CONSOLIDADO)
    arquivos = sorted(
        f for f in glob.glob(os.path.join(pasta_bases, '*.xlsx'))
        if os.path.basename(f) != NOME_CONSOLIDADO
    )

    if not arquivos:
        return

    wb_consolidado = carregar_consolidado(caminho_consolidado)
    total_inseridos = {ABA_CURITIBA: 0, ABA_TOLEDO: 0}

    for arquivo in arquivos:
        unidade_texto, headers, rows = _ler_individual(arquivo)
        nome_aba = obter_unidade(unidade_texto)
        if nome_aba is None:
            print(f"  [AVISO] Unidade não identificada em {os.path.basename(arquivo)} — ignorado na consolidação")
            continue

        ws = wb_consolidado[nome_aba]
        cabecalho = _garantir_cabecalho(ws, headers)
        total_inseridos[nome_aba] += inserir_registros(ws, cabecalho, headers, rows)

    for nome_aba in (ABA_CURITIBA, ABA_TOLEDO):
        ws = wb_consolidado[nome_aba]
        if ws.max_row >= 2 and ws.max_column >= 1:
            ultima_coluna = ws.cell(row=1, column=ws.max_column).column_letter
            ws.auto_filter.ref = f"A1:{ultima_coluna}{ws.max_row}"

    wb_consolidado.save(caminho_consolidado)
    print(f"\nConsolidado atualizado: +{total_inseridos[ABA_CURITIBA]} CURITIBA, "
          f"+{total_inseridos[ABA_TOLEDO]} TOLEDO ({NOME_CONSOLIDADO})")


if __name__ == '__main__':
    pasta = os.path.abspath(PASTA_BASES)

    if not os.path.isdir(pasta):
        print(f"[ERRO] Pasta não encontrada: {pasta}")
        exit(1)

    csvs = glob.glob(os.path.join(pasta, '*.csv'))

    if csvs:
        print(f"Encontrados {len(csvs)} CSV(s) em: {pasta}\n")
        for arquivo in sorted(csvs):
            processar(arquivo)
    else:
        print(f"Nenhum CSV encontrado em: {pasta}")

    # Bases que já chegam em XLSX (ex.: relatórios de cancela/catraca) também
    # são padronizadas; arquivos já estilizados por este script são ignorados.
    xlsx_brutos = [
        f for f in glob.glob(os.path.join(pasta, '*.xlsx'))
        if os.path.basename(f) != NOME_CONSOLIDADO
    ]
    for arquivo in sorted(xlsx_brutos):
        processar_xlsx_bruto(arquivo)

    # A consolidação roda sempre, mesmo sem CSV/XLSX novo, para absorver bases já
    # presentes em bases/ que ainda não tenham sido incorporadas ao histórico.
    consolidar_planilhas(pasta)

    print("\nConcluído.")